#!/usr/bin/env python3
"""Pretvara prezentacijske tablice G1, G2, G3, G5 i G6 u jednu CSV bazu.

Zrno izlazne baze jest:
    datum x izvorna tablica x ekonomska serija x vrsta kamatne stope

Svaki redak sadrzi dvije uparene mjere: kamatnu stopu i pripadajuci iznos.
CSV je prilagodjen hrvatskom Excelu: UTF-8 BOM, tocka-zarez kao delimiter i
decimalni zarez. Potpuno prazna opazanja standardno se izostavljaju, dok se
izvorni znak "–" cuva u statusu mjere.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


UKUPNO = "UKUPNO"
NP = "NIJE_PRIMJENJIVO"
NOVI = "NOVI_POSLOVI_TIJEKOM_MJESECA"
STANJE = "STANJE_NA_KRAJU_MJESECA"


@dataclass(frozen=True)
class SeriesMeta:
    pozicija: str
    sektor: str = UKUPNO
    podskup: str = UKUPNO
    instrument: str = UKUPNO
    namjena: str = NP
    dospijece: str = NP
    otkazni_rok: str = NP
    fiksiranje: str = NP
    velicina: str = NP
    osnova: str = NOVI
    fusnota_a: bool = False


@dataclass(frozen=True)
class BlockSpec:
    sheet: str
    vrsta_stope: str
    koncept_tablice: str
    date_row_rate: int
    date_row_amount: int
    status_row_rate: int
    status_row_amount: int
    amount_offset: int
    rows: dict[int, SeriesMeta]


def dep(
    sektor: str = UKUPNO,
    instrument: str = UKUPNO,
    *,
    dospijece: str = NP,
    otkazni_rok: str = NP,
    osnova: str = NOVI,
    fusnota_a: bool = False,
) -> SeriesMeta:
    return SeriesMeta(
        pozicija="DEPOZITI",
        sektor=sektor,
        instrument=instrument,
        namjena=NP,
        dospijece=dospijece,
        otkazni_rok=otkazni_rok,
        fiksiranje=NP,
        velicina=NP,
        osnova=osnova,
        fusnota_a=fusnota_a,
    )


def loan(
    sektor: str,
    *,
    podskup: str = UKUPNO,
    instrument: str = "KREDITI",
    namjena: str = UKUPNO,
    dospijece: str = UKUPNO,
    fiksiranje: str = UKUPNO,
    velicina: str = UKUPNO,
    osnova: str = NOVI,
    fusnota_a: bool = False,
) -> SeriesMeta:
    return SeriesMeta(
        pozicija="KREDITI",
        sektor=sektor,
        podskup=podskup,
        instrument=instrument,
        namjena=namjena,
        dospijece=dospijece,
        otkazni_rok=NP,
        fiksiranje=fiksiranje,
        velicina=velicina,
        osnova=osnova,
        fusnota_a=fusnota_a,
    )


def g1_rows() -> dict[int, SeriesMeta]:
    h, n = "KUCANSTVA", "NEFINANCIJSKA_DRUSTVA"
    r: dict[int, SeriesMeta] = {
        10: dep(h, "PREKONOCNI_DEPOZITI", osnova=STANJE),
        11: dep(h, "TRANSAKCIJSKI_RACUNI", osnova=STANJE),
        12: dep(h, "STEDNI_DEPOZITI", osnova=STANJE),
        13: dep(h, "OROCENI_DEPOZITI", dospijece=UKUPNO, fusnota_a=True),
        14: dep(h, "OROCENI_DEPOZITI", dospijece="DO_3_MJESECA"),
        15: dep(h, "OROCENI_DEPOZITI", dospijece="OD_3_DO_6_MJESECI"),
        16: dep(h, "OROCENI_DEPOZITI", dospijece="OD_6_MJESECI_DO_1_GODINE"),
        17: dep(h, "OROCENI_DEPOZITI", dospijece="OD_1_DO_2_GODINE"),
        18: dep(h, "OROCENI_DEPOZITI", dospijece="VISE_OD_2_GODINE"),
        19: dep(h, "DEPOZITI_U_OTKAZNOM_ROKU", otkazni_rok=UKUPNO),
        20: dep(h, "DEPOZITI_U_OTKAZNOM_ROKU", otkazni_rok="DO_3_MJESECA"),
        21: dep(h, "DEPOZITI_U_OTKAZNOM_ROKU", otkazni_rok="VISE_OD_3_MJESECA"),
        23: dep(n, "PREKONOCNI_DEPOZITI", osnova=STANJE),
        24: dep(n, "TRANSAKCIJSKI_RACUNI", osnova=STANJE),
        25: dep(n, "STEDNI_DEPOZITI", osnova=STANJE),
        26: dep(n, "OROCENI_DEPOZITI", dospijece=UKUPNO, fusnota_a=True),
        27: dep(n, "OROCENI_DEPOZITI", dospijece="DO_3_MJESECA"),
        28: dep(n, "OROCENI_DEPOZITI", dospijece="OD_3_DO_6_MJESECI"),
        29: dep(n, "OROCENI_DEPOZITI", dospijece="OD_6_MJESECI_DO_1_GODINE"),
        30: dep(n, "OROCENI_DEPOZITI", dospijece="OD_1_DO_2_GODINE"),
        31: dep(n, "OROCENI_DEPOZITI", dospijece="VISE_OD_2_GODINE"),
        32: dep(UKUPNO, "REPO_POSLOVI", dospijece=UKUPNO),
    }
    return r


def g2_nominal_rows() -> dict[int, SeriesMeta]:
    h = "KUCANSTVA"
    other = "OSTALE_NAMJENE"
    cash = "GOTOVINSKI_NENAMJENSKI"
    r: dict[int, SeriesMeta] = {
        9: loan(h, instrument="REVOLVING_PREKORACENJA_I_KARTICE", dospijece=NP, fiksiranje=NP, osnova=STANJE),
        10: loan(h, instrument="REVOLVING_KREDITI", dospijece=NP, fiksiranje=NP, osnova=STANJE),
        11: loan(h, instrument="PREKORACENJA_PO_TRANSAKCIJSKOM_RACUNU", dospijece=NP, fiksiranje=NP, osnova=STANJE),
        12: loan(h, instrument="KREDITI_PO_KREDITNIM_KARTICAMA", dospijece=NP, fiksiranje=NP, osnova=STANJE),
        13: loan(h, podskup="OBRTNICI", instrument="REVOLVING_PREKORACENJA_I_KARTICE", dospijece=NP, fiksiranje=NP, osnova=STANJE),
        14: loan(h, namjena="POTROSACKI", fusnota_a=True),
        15: loan(h, namjena="POTROSACKI", fiksiranje="PROMJENJIVA_ILI_DO_1_GODINE"),
        16: loan(h, namjena="POTROSACKI", fiksiranje="OD_1_DO_5_GODINA"),
        17: loan(h, namjena="POTROSACKI", fiksiranje="VISE_OD_5_GODINA"),
        18: loan(h, namjena="POTROSACKI", fiksiranje="FIKSNA_CIJELI_VIJEK"),
        19: loan(h, namjena="STAMBENI", fusnota_a=True),
        20: loan(h, namjena="STAMBENI", fiksiranje="PROMJENJIVA_ILI_DO_1_GODINE"),
        21: loan(h, namjena="STAMBENI", fiksiranje="OD_1_DO_5_GODINA"),
        22: loan(h, namjena="STAMBENI", fiksiranje="OD_5_DO_10_GODINA"),
        23: loan(h, namjena="STAMBENI", fiksiranje="VISE_OD_10_GODINA"),
        24: loan(h, namjena="STAMBENI", fiksiranje="FIKSNA_CIJELI_VIJEK"),
        25: loan(h, namjena="STAMBENI", dospijece="KRATKOROCNO", fiksiranje="FIKSNA_CIJELI_VIJEK"),
        26: loan(h, namjena="STAMBENI", dospijece="DUGOROCNO", fiksiranje="FIKSNA_CIJELI_VIJEK"),
        27: loan(h, namjena=other, fusnota_a=True),
        28: loan(h, namjena=other, fiksiranje="PROMJENJIVA_ILI_DO_1_GODINE"),
        29: loan(h, namjena=other, fiksiranje="OD_1_DO_5_GODINA"),
        30: loan(h, namjena=other, fiksiranje="VISE_OD_5_GODINA"),
        31: loan(h, namjena=cash),
        32: loan(h, namjena=cash, fiksiranje="PROMJENJIVA_ILI_DO_1_GODINE"),
        33: loan(h, namjena=cash, fiksiranje="OD_1_DO_5_GODINA"),
        34: loan(h, namjena=cash, fiksiranje="VISE_OD_5_GODINA"),
        35: loan(h, namjena=cash, fiksiranje="FIKSNA_CIJELI_VIJEK"),
        36: loan(h, podskup="OBRTNICI", namjena=other),
    }
    return r


def g2_effective_rows() -> dict[int, SeriesMeta]:
    h = "KUCANSTVA"
    return {
        75: loan(h, namjena="POTROSACKI", fusnota_a=True),
        76: loan(h, namjena="STAMBENI", fusnota_a=True),
        77: loan(h, namjena="OSTALE_NAMJENE"),
        78: loan(h, namjena="GOTOVINSKI_NENAMJENSKI"),
    }


def g3_rows() -> dict[int, SeriesMeta]:
    n = "NEFINANCIJSKA_DRUSTVA"
    r: dict[int, SeriesMeta] = {
        9: loan(n, instrument="REVOLVING_PREKORACENJA_I_KARTICE", dospijece=NP, fiksiranje=NP, velicina=NP, osnova=STANJE),
        10: loan(n, instrument="REVOLVING_I_PREKORACENJA", dospijece=NP, fiksiranje=NP, velicina=NP, osnova=STANJE),
        11: loan(n, instrument="KREDITI_PO_KREDITNIM_KARTICAMA", dospijece=NP, fiksiranje=NP, velicina=NP, osnova=STANJE),
    }
    fixation_rows = {
        1: "PROMJENJIVA_ILI_DO_3_MJESECA",
        2: "OD_3_MJESECA_DO_1_GODINE",
        3: "OD_1_DO_3_GODINE",
        4: "OD_3_DO_5_GODINA",
        5: "OD_5_DO_10_GODINA",
        6: "VISE_OD_10_GODINA",
    }
    for base, size in ((12, "DO_0_25_MIL_EUR"), (19, "OD_0_25_DO_1_MIL_EUR"), (26, "VISE_OD_1_MIL_EUR")):
        r[base] = loan(n, velicina=size, fusnota_a=True)
        for offset, fixation in fixation_rows.items():
            r[base + offset] = loan(n, velicina=size, fiksiranje=fixation)
    return r


def g5_rows() -> dict[int, SeriesMeta]:
    h, n = "KUCANSTVA", "NEFINANCIJSKA_DRUSTVA"
    r: dict[int, SeriesMeta] = {
        9: dep(UKUPNO, UKUPNO, dospijece=UKUPNO, osnova=STANJE),
        10: dep(UKUPNO, "OROCENI_DEPOZITI", dospijece=UKUPNO, osnova=STANJE, fusnota_a=True),
        11: dep(h, "OROCENI_DEPOZITI", dospijece=UKUPNO, osnova=STANJE),
        12: dep(h, "OROCENI_DEPOZITI", dospijece="KRATKOROCNO", osnova=STANJE),
        13: dep(h, "OROCENI_DEPOZITI", dospijece="DO_3_MJESECA", osnova=STANJE),
        14: dep(h, "OROCENI_DEPOZITI", dospijece="OD_3_DO_6_MJESECI", osnova=STANJE),
        15: dep(h, "OROCENI_DEPOZITI", dospijece="OD_6_MJESECI_DO_1_GODINE", osnova=STANJE),
        16: dep(h, "OROCENI_DEPOZITI", dospijece="DUGOROCNO", osnova=STANJE),
        17: dep(h, "OROCENI_DEPOZITI", dospijece="OD_1_DO_2_GODINE", osnova=STANJE),
        18: dep(h, "OROCENI_DEPOZITI", dospijece="VISE_OD_2_GODINE", osnova=STANJE),
        19: dep(n, "OROCENI_DEPOZITI", dospijece=UKUPNO, osnova=STANJE),
        20: dep(n, "OROCENI_DEPOZITI", dospijece="KRATKOROCNO", osnova=STANJE),
        21: dep(n, "OROCENI_DEPOZITI", dospijece="DO_3_MJESECA", osnova=STANJE),
        22: dep(n, "OROCENI_DEPOZITI", dospijece="OD_3_DO_6_MJESECI", osnova=STANJE),
        23: dep(n, "OROCENI_DEPOZITI", dospijece="OD_6_MJESECI_DO_1_GODINE", osnova=STANJE),
        24: dep(n, "OROCENI_DEPOZITI", dospijece="DUGOROCNO", osnova=STANJE),
        25: dep(n, "OROCENI_DEPOZITI", dospijece="OD_1_DO_2_GODINE", osnova=STANJE),
        26: dep(n, "OROCENI_DEPOZITI", dospijece="VISE_OD_2_GODINE", osnova=STANJE),
        27: dep(UKUPNO, "REPO_POSLOVI", dospijece=UKUPNO, osnova=STANJE),
        28: loan(UKUPNO, osnova=STANJE),
        29: loan(h, osnova=STANJE),
        30: loan(h, namjena="STAMBENI", osnova=STANJE, fusnota_a=True),
        31: loan(h, namjena="STAMBENI", dospijece="KRATKOROCNO", osnova=STANJE),
        32: loan(h, namjena="STAMBENI", dospijece="DUGOROCNO", osnova=STANJE),
        33: loan(h, namjena="STAMBENI", dospijece="OD_1_DO_5_GODINA", osnova=STANJE),
        34: loan(h, namjena="STAMBENI", dospijece="VISE_OD_5_GODINA", osnova=STANJE),
        35: loan(h, namjena="POTROSACKI_I_OSTALI", osnova=STANJE, fusnota_a=True),
        36: loan(h, namjena="POTROSACKI_I_OSTALI", dospijece="KRATKOROCNO", osnova=STANJE, fusnota_a=True),
        37: loan(h, namjena="POTROSACKI_I_OSTALI", dospijece="DUGOROCNO", osnova=STANJE, fusnota_a=True),
        38: loan(h, namjena="POTROSACKI_I_OSTALI", dospijece="OD_1_DO_5_GODINA", osnova=STANJE),
        39: loan(h, namjena="POTROSACKI_I_OSTALI", dospijece="VISE_OD_5_GODINA", osnova=STANJE),
        40: loan(h, podskup="OBRTNICI", namjena="POTROSACKI_I_OSTALI", osnova=STANJE),
        41: loan(n, osnova=STANJE, fusnota_a=True),
        42: loan(n, dospijece="KRATKOROCNO", osnova=STANJE),
        43: loan(n, dospijece="DUGOROCNO", osnova=STANJE),
        44: loan(n, dospijece="OD_1_DO_5_GODINA", osnova=STANJE),
        45: loan(n, dospijece="VISE_OD_5_GODINA", osnova=STANJE),
    }
    return r


def g6_rows() -> dict[int, SeriesMeta]:
    h, n = "KUCANSTVA", "NEFINANCIJSKA_DRUSTVA"
    r: dict[int, SeriesMeta] = {
        9: dep(UKUPNO, UKUPNO, dospijece=UKUPNO),
        10: dep(UKUPNO, "OROCENI_DEPOZITI", dospijece=UKUPNO, fusnota_a=True),
        11: dep(h, "OROCENI_DEPOZITI", dospijece=UKUPNO),
        12: dep(h, "OROCENI_DEPOZITI", dospijece="KRATKOROCNO"),
        13: dep(h, "OROCENI_DEPOZITI", dospijece="DUGOROCNO"),
        14: dep(n, "OROCENI_DEPOZITI", dospijece=UKUPNO),
        15: dep(n, "OROCENI_DEPOZITI", dospijece="KRATKOROCNO"),
        16: dep(n, "OROCENI_DEPOZITI", dospijece="DUGOROCNO"),
        17: dep(UKUPNO, "REPO_POSLOVI", dospijece=UKUPNO),
        18: loan(UKUPNO),
        19: loan(h),
        20: loan(h, namjena="STAMBENI", fusnota_a=True),
        21: loan(h, namjena="STAMBENI", dospijece="KRATKOROCNO"),
        22: loan(h, namjena="STAMBENI", dospijece="DUGOROCNO"),
        23: loan(h, namjena="STAMBENI", dospijece="OD_1_DO_5_GODINA"),
        24: loan(h, namjena="STAMBENI", dospijece="VISE_OD_5_GODINA"),
        25: loan(h, namjena="POTROSACKI_I_OSTALI", fusnota_a=True),
        26: loan(h, namjena="POTROSACKI_I_OSTALI", dospijece="KRATKOROCNO", fusnota_a=True),
        27: loan(h, namjena="POTROSACKI_I_OSTALI", dospijece="DUGOROCNO", fusnota_a=True),
        28: loan(h, namjena="POTROSACKI_I_OSTALI", dospijece="OD_1_DO_5_GODINA"),
        29: loan(h, namjena="POTROSACKI_I_OSTALI", dospijece="VISE_OD_5_GODINA"),
        30: loan(h, podskup="OBRTNICI", namjena="POTROSACKI_I_OSTALI"),
        31: loan(n, fusnota_a=True),
        32: loan(n, dospijece="KRATKOROCNO"),
        33: loan(n, dospijece="DUGOROCNO"),
        34: loan(n, dospijece="OD_1_DO_5_GODINA"),
        35: loan(n, dospijece="VISE_OD_5_GODINA"),
    }
    return r


BLOCKS = (
    BlockSpec("G1", "NOMINALNA", "NOVI_POSLOVI", 8, 37, 7, 36, 29, g1_rows()),
    BlockSpec("G2", "NOMINALNA", "NOVI_POSLOVI", 8, 41, 7, 40, 33, g2_nominal_rows()),
    BlockSpec("G2", "EFEKTIVNA", "NOVI_POSLOVI", 74, 83, 73, 82, 9, g2_effective_rows()),
    BlockSpec("G3", "NOMINALNA", "NOVI_POSLOVI", 8, 37, 7, 36, 29, g3_rows()),
    BlockSpec("G5", "NOMINALNA", "STANJA", 8, 50, 7, 49, 42, g5_rows()),
    BlockSpec("G6", "NOMINALNA", "NOVI_POSLOVI", 8, 40, 7, 39, 32, g6_rows()),
)


FOOTNOTE_TEXT = {
    "G1": "Povijesni podaci u ovom retku odnose se samo na depozite u eurima i kunama s valutnom klauzulom uz euro.",
    "G2": "Povijesni podaci u ovom retku odnose se samo na kredite u eurima i kunama s valutnom klauzulom uz euro.",
    "G3": "Povijesni podaci u ovom retku odnose se samo na kredite u eurima i kunama s valutnom klauzulom uz euro.",
    "G5": "Povijesni podaci u ovom retku odnose se samo na kredite odnosno depozite u eurima i kunama bez valutne klauzule te s valutnom klauzulom uz euro.",
    "G6": "Povijesni podaci u ovom retku odnose se samo na kredite odnosno depozite u eurima i kunama s valutnom klauzulom uz euro.",
}

EURO_NOTE = {
    "G1": "Počevši sa siječnjem 2023. svi podaci odnose se samo na depozite u eurima.",
    "G2": "Počevši sa siječnjem 2023. svi podaci odnose se samo na kredite u eurima.",
    "G3": "Počevši sa siječnjem 2023. svi podaci odnose se samo na kredite u eurima.",
    "G5": "Počevši sa siječnjem 2023. svi podaci odnose se samo na kredite i depozite u eurima.",
    "G6": "Počevši sa siječnjem 2023. svi podaci odnose se samo na kredite i depozite u eurima.",
}

FIELDS = (
    "datum",
    "tablica",
    "serija_id",
    "koncept_tablice",
    "osnova_obracuna",
    "pozicija",
    "sektor_protustranke",
    "podskup_protustranke",
    "instrument",
    "namjena_kredita",
    "izvorno_dospijece",
    "otkazni_rok",
    "pocetno_fiksiranje_kamatne_stope",
    "velicina_kredita",
    "vrsta_kamatne_stope",
    "kamatna_stopa",
    "iznos",
    "status_kamatne_stope",
    "status_iznosa",
    "valutni_obuhvat",
    "frekvencija",
    "jedinica_kamatne_stope",
    "jedinica_iznosa",
    "broj_decimala_stope",
    "broj_decimala_iznosa",
    "izvorna_oznaka",
    "hijerarhijska_oznaka",
    "nadredena_serija_id",
    "od_toga",
    "izvorni_redak_stope",
    "izvorni_redak_iznosa",
    "napomena_serije",
)


def normalize_space(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ").replace("\t", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def hierarchy_code(raw_label: str) -> str:
    match = re.match(r"^\s*(\d+(?:\.\d+)*\.)", raw_label.replace("\xa0", " "))
    return match.group(1) if match else ""


def clean_label(raw_label: str, footnote_a: bool) -> str:
    label = normalize_space(raw_label)
    label = re.sub(r"^\d+(?:\.\d+)*\.\s*", "", label)
    if footnote_a and label.endswith("a"):
        label = label[:-1].rstrip()
    return label


def indent_width(raw_label: str) -> int:
    expanded = raw_label.replace("\xa0", " ").expandtabs(4)
    return len(expanded) - len(expanded.lstrip(" "))


def series_id(block: BlockSpec, rate_row: int) -> str:
    rate_code = "NKS" if block.vrsta_stope == "NOMINALNA" else "EKS"
    return f"{block.sheet}_{rate_code}_R{rate_row:03d}"


def parent_ids(block: BlockSpec, ws: Any) -> dict[int, str]:
    stack: list[tuple[int, str]] = []
    result: dict[int, str] = {}
    for row in sorted(block.rows):
        raw = str(ws.cell(row, 2).value or "")
        indent = indent_width(raw)
        while stack and stack[-1][0] >= indent:
            stack.pop()
        result[row] = stack[-1][1] if stack else ""
        stack.append((indent, series_id(block, row)))
    return result


def cell_kind(value: Any) -> str:
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return "NUMERIC"
    if value == "–":
        return "DASH"
    if value in (None, ""):
        return "BLANK"
    raise ValueError(f"Neocekivana vrijednost u podatkovnoj celiji: {value!r}")


def csv_number(value: Any) -> str:
    if cell_kind(value) != "NUMERIC":
        return ""
    # Excel racuna s najvise 15 znacajnih znamenaka. Time se cuva puna
    # smislena preciznost, ali se ne prenose binarni artefakti poput
    # 1.6201416000000002 umjesto 1.6201416.
    text = format(value, ".15g") if isinstance(value, float) else format(Decimal(str(value)), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text.replace(".", ",")


def measure_status(value: Any, marker: Any) -> str:
    kind = cell_kind(value)
    if kind == "DASH":
        return "IZVORNI_ZNAK_MINUS"
    if kind == "BLANK":
        return "NEMA_PODATKA"
    return "REVIDIRANO" if marker == "*" else "REDOVNO"


def currency_scope(sheet: str, meta: SeriesMeta, date: datetime) -> str:
    if date >= datetime(2023, 1, 1):
        return "EUR"
    if not meta.fusnota_a:
        return "PREMA_OPCEM_POVIJESNOM_OBUHVATU_TABLICE"
    if sheet == "G5":
        return "EUR_I_HRK_BEZ_KLAUZULE_TE_HRK_UZ_EUR_KLAUZULU"
    return "EUR_I_HRK_UZ_EUR_KLAUZULU"


def note_for(sheet: str, meta: SeriesMeta) -> str:
    parts = [EURO_NOTE[sheet]]
    if meta.fusnota_a:
        parts.insert(0, FOOTNOTE_TEXT[sheet])
    return " ".join(parts)


def date_columns(ws: Any, row: int) -> list[int]:
    cols = [c for c in range(3, ws.max_column + 1) if isinstance(ws.cell(row, c).value, datetime)]
    if not cols:
        raise ValueError(f"{ws.title}: nisu pronadjeni datumi u retku {row}.")
    expected = list(range(cols[0], cols[-1] + 1))
    if cols != expected:
        raise ValueError(f"{ws.title}: datumski stupci u retku {row} nisu neprekinuti.")
    return cols


def validate_monthly(dates: Iterable[datetime], context: str) -> None:
    values = list(dates)
    for previous, current in zip(values, values[1:]):
        distance = (current.year - previous.year) * 12 + current.month - previous.month
        if distance != 1:
            raise ValueError(f"{context}: datumi nisu mjesecno neprekinuti: {previous} -> {current}")


def build_rows(input_path: Path, keep_blank: bool = False) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(input_path, data_only=True, read_only=False)
    required = {"Metodologija", "G1", "G2", "G3", "G5", "G6"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"Nedostaju ocekivani listovi: {sorted(missing)}")

    output: list[dict[str, Any]] = []
    seen_series: set[str] = set()
    possible = numeric = dash = blank = 0

    for block in BLOCKS:
        ws = workbook[block.sheet]
        rate_cols = date_columns(ws, block.date_row_rate)
        amount_cols = date_columns(ws, block.date_row_amount)
        if rate_cols != amount_cols:
            raise ValueError(f"{block.sheet} {block.vrsta_stope}: datumski stupci stopa i iznosa nisu jednaki.")

        rate_dates = [ws.cell(block.date_row_rate, c).value for c in rate_cols]
        amount_dates = [ws.cell(block.date_row_amount, c).value for c in amount_cols]
        if rate_dates != amount_dates:
            raise ValueError(f"{block.sheet} {block.vrsta_stope}: datumi stopa i iznosa nisu jednaki.")
        validate_monthly(rate_dates, f"{block.sheet} {block.vrsta_stope}")

        parents = parent_ids(block, ws)
        for rate_row, meta in sorted(block.rows.items()):
            amount_row = rate_row + block.amount_offset
            raw_rate_label = str(ws.cell(rate_row, 2).value or "")
            raw_amount_label = str(ws.cell(amount_row, 2).value or "")
            if normalize_space(raw_rate_label) != normalize_space(raw_amount_label):
                raise ValueError(
                    f"{block.sheet}: oznake retka stope {rate_row} i iznosa {amount_row} nisu jednake."
                )

            sid = series_id(block, rate_row)
            if sid in seen_series:
                raise ValueError(f"Duplicirani serija_id: {sid}")
            seen_series.add(sid)

            cleaned = clean_label(raw_rate_label, meta.fusnota_a)
            code = hierarchy_code(raw_rate_label)
            is_of_which = cleaned.casefold().startswith("od toga:")

            for col, date in zip(rate_cols, rate_dates):
                possible += 1
                rate_value = ws.cell(rate_row, col).value
                amount_value = ws.cell(amount_row, col).value
                rate_kind = cell_kind(rate_value)
                amount_kind = cell_kind(amount_value)
                if rate_kind != amount_kind:
                    raise ValueError(
                        f"{block.sheet} {sid} {date:%Y-%m-%d}: raspolozivost stope i iznosa nije jednaka."
                    )
                if rate_kind == "BLANK":
                    blank += 1
                    if not keep_blank:
                        continue
                elif rate_kind == "DASH":
                    dash += 1
                else:
                    numeric += 1

                output.append(
                    {
                        "datum": date.strftime("%Y-%m-%d"),
                        "tablica": block.sheet,
                        "serija_id": sid,
                        "koncept_tablice": block.koncept_tablice,
                        "osnova_obracuna": meta.osnova,
                        "pozicija": meta.pozicija,
                        "sektor_protustranke": meta.sektor,
                        "podskup_protustranke": meta.podskup,
                        "instrument": meta.instrument,
                        "namjena_kredita": meta.namjena,
                        "izvorno_dospijece": meta.dospijece,
                        "otkazni_rok": meta.otkazni_rok,
                        "pocetno_fiksiranje_kamatne_stope": meta.fiksiranje,
                        "velicina_kredita": meta.velicina,
                        "vrsta_kamatne_stope": block.vrsta_stope,
                        "kamatna_stopa": csv_number(rate_value),
                        "iznos": csv_number(amount_value),
                        "status_kamatne_stope": measure_status(
                            rate_value, ws.cell(block.status_row_rate, col).value
                        ),
                        "status_iznosa": measure_status(
                            amount_value, ws.cell(block.status_row_amount, col).value
                        ),
                        "valutni_obuhvat": currency_scope(block.sheet, meta, date),
                        "frekvencija": "M",
                        "jedinica_kamatne_stope": "POSTOTAK_GODISNJE",
                        "jedinica_iznosa": "MIL_EUR",
                        "broj_decimala_stope": "2",
                        "broj_decimala_iznosa": "1",
                        "izvorna_oznaka": cleaned,
                        "hijerarhijska_oznaka": code,
                        "nadredena_serija_id": parents[rate_row],
                        "od_toga": "DA" if is_of_which else "NE",
                        "izvorni_redak_stope": str(rate_row),
                        "izvorni_redak_iznosa": str(amount_row),
                        "napomena_serije": note_for(block.sheet, meta),
                    }
                )

    summary = {
        "input": str(input_path),
        "broj_serija": len(seen_series),
        "broj_mogucih_redaka": possible,
        "broj_numerickih_redaka": numeric,
        "broj_redaka_sa_znakom_minus": dash,
        "broj_potpuno_praznih_redaka": blank,
        "broj_izvezenih_redaka": len(output),
        "ukljuceni_potpuno_prazni_retci": keep_blank,
    }
    return output, summary


def write_csv(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=FIELDS,
            delimiter=";",
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Ulazna Excel knjiga G tablice.xlsx")
    parser.add_argument("output", type=Path, nargs="?", default=Path("G_podaci.csv"), help="Izlazni CSV")
    parser.add_argument(
        "--keep-blank",
        action="store_true",
        help="Uključi i potpuno prazne kombinacije serije i datuma.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows, summary = build_rows(args.input, keep_blank=args.keep_blank)
    if summary["broj_serija"] != 142:
        raise ValueError(f"Ocekivane su 142 serije, pronadjeno je {summary['broj_serija']}.")
    write_csv(rows, args.output)
    summary["output"] = str(args.output)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
